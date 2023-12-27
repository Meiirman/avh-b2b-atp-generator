import os
import traceback
from openpyxl import load_workbook
from shutil import copyfile

import pandas as pd
import locale
locale.setlocale(locale.LC_ALL, '')


def generate(render_data, template_path, output_folder_path) -> dict:
    if render_data is None:
        return {"status" : "error", "message" : "Вызов метода не может быть пустым"}
    return render_and_save_excel(render_data, template_path, output_folder_path)

def render_and_save_excel(render_data, template_path, output_folder_path):
    print(render_data)

    render_data["data"]["clent"] = render_data["data"]["BS_COMPANY"]
    render_data["data"]["city"] = render_data["data"]["ORDER_REGION"]
    render_data["data"]["address"] = 	render_data["data"]["BS_ADDRESS"]
    render_data["data"]["number"] = render_data["data"]["BS_NUMBER"]
    
    try: render_data["data"]["date"] = render_data["data"]["ORDER_DOGOVOR_DATE"]
    except:
        traceback.print_exc()
        input()
        render_data["data"]["date"] = 0 

    
    try: render_data["data"]["budget"] =  render_data["data"]["TOTAL_NDS"]
    except:
        traceback.print_exc()
        input()
        render_data["data"]["budget"] = 0 

    try: render_data["data"]["budget_with_nds"] =  render_data["data"]["TOTAL_SUMM_NDS"]
    except:
        traceback.print_exc()
        input()
        render_data["data"]["budget_with_nds"] = 0 

    try: render_data["data"]["nds_budget"] = float(render_data["data"]["budget_with_nds"]) - float(render_data["data"]["budget"])
    except:
        pass

    list_table = render_data["data"]["TABLE"]
    
    if list_table:
        try:
            for i, row in enumerate(list_table):
                if i == 0 :
                    continue

                render_data["data"]["index_" + str(i)] = row["index"]
                render_data["data"]["table_number_" + str(i)] = row["number"]
                render_data["data"]["table_work_name_" + str(i)] = row["work_name"]
                render_data["data"]["table_measure_" + str(i)] = row["measure"]
                render_data["data"]["table_count_" + str(i)] = row["count"]
                render_data["data"]["table_price_" + str(i)] = row["price"]
                render_data["data"]["table_price_with_nds_" + str(i)] = row["price_with_nds"] 

            # Копируем шаблон в выходную директорию
            output_file_path = os.path.join(output_folder_path, "output.xlsx")
            copyfile(template_path, output_file_path)

            # Переименовываем копию в "output.xlsx"
            os.rename(output_file_path, os.path.join(output_folder_path, "output.xlsx"))

            # Загружаем данные из созданной копии
            wb = load_workbook(os.path.join(output_folder_path, "output.xlsx"))
            ws = wb.active

            # Проходимся по каждой ячейке в созданной копии
            for i, row in enumerate(ws.iter_rows(min_row=1, max_row=150, min_col=1, max_col=30)):
                for cell in row:
                    for key, value in render_data["data"].items():
                        try:str_cell_value = str(cell.value)
                        except: str_cell_value = "None"

                        if "{{"+f"table_count_{i}"+"}}" in key and "{{" + f"table_count_{i}" + "}}" in str_cell_value:
                            try: cell.value = cell.value.replace("{{" + key + "}}", str(value).replace(".", ","))
                            except: pass

                        elif "{{"+f"table_price_{i}"+"}}" in key and "{{"+f"table_price_{i}"+"}}" + "}}" in str_cell_value:
                            try:
                                cell.value = str(value).replace(".", ",")
                                # cell.value = value
                                cell.number_format = '$# ##0.00'  
                            except: 
                                pass

                        elif "{{"+f"table_price_with_nds_{i}"+"}}" in key and "{{"+f"table_price_with_nds_{i}"+"}}" in str_cell_value:
                            try:
                                cell.value = str(value).replace(".", ",")
                                # cell.value = value
                                cell.number_format = '$# ##0.00'  
                            except: 
                                pass

                        elif "00:00:00" in str(value) and cell.value and "{{date}}" in str(cell.value):
                            try: 
                                from datetime import datetime
                                months = {
                                    "01" : " января ",
                                    "02" : " февраля ",
                                    "03" : " марта ",
                                    "04" : " апреля ",
                                    "05" : " мая ",
                                    "06" : " июня ",
                                    "07" : " июля ",
                                    "08" : " августа ",
                                    "09" : " сентября ",
                                    "10" : " октября ",
                                    "11" : " ноября ",
                                    "12" : " декабря "
                                }
                                
                                date_string = str(value)
                                date_object = datetime.strptime(date_string, "%Y-%m-%d %H:%M:%S")
                                formatted_date = date_object.strftime("%d.%m.%Y")
                                dmy = str(formatted_date.split(".")[0])
                                dmy += str(months[formatted_date.split(".")[1]])
                                dmy += str(formatted_date.split(".")[2])
                                formatted_date = dmy
                                # cell.value = "от " + formatted_date
                                cell.value = cell.value.replace("{{" + key + "}}", formatted_date)
                            except: 
                                traceback.print_exc()
                        
                        elif "price" in key and "price" in str(cell.value).lower() and f"{i}" + "}}" in str(cell.value).lower():
                            try:
                                cell.value = str(value).replace(".", ",")
                                cell.number_format = '# ##0.00' 
                            except: 
                                traceback.print_exc()

                        else:                        
                            try: 
                                try:
                                    cell.value = cell.value.replace("{{" + key + "}}", str(value)).replace(".", ",")
                                except:
                                    cell.value = cell.value.replace("{{" + key + "}}", str(value))

                            except: pass
                        


            del_rows = []
            del_rows_nums = []
            for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=150, min_col=1, max_col=30)):
                row_text = " ".join([str(cell.value) for cell in row])
                if "{{" in row_text and "}}" in row_text and "index" in row_text:
                    del_rows_nums.append(idx)
                    del_rows.append(row)

            for row in del_rows:
                for cell in row:
                    cell.value = None

            ws.delete_rows(14+len(list_table), amount=len(del_rows))    

            # ws[f'A{16+len(list_table)}:Q{15+len(list_table)+30}']
            start_row = 14+len(list_table)
            

            fwea = []
            for row in list_table:
                try: fwea.append(float(row["price"])*float(row["count"]))
                except: 
                    print("\n__________"*10)
                    traceback.print_exc()
                    print("\n__________"*10)


            locale.setlocale(locale.LC_ALL, '')

            TOTAL_SUM = locale.format_string("%0.2f", round(sum(fwea), 2), grouping=True)
            TOTAL_NDS = locale.format_string("%0.2f", round(sum(fwea)*0.12, 2), grouping=True)
            TOTAL_SUM_NDS = locale.format_string("%0.2f", round(sum(fwea)*1.12, 2), grouping=True)



            ws[f'L{start_row}'] = f'{TOTAL_SUM}'
            ws[f'L{start_row+1}'] = f'{TOTAL_SUM_NDS}'

            ws[f'B11'] = str(ws[f'B11'].value).replace('L80', f'{TOTAL_NDS}').replace('L79', f'{TOTAL_SUM}')

            ws.merge_cells(f'E{start_row+0}:K{start_row+0}')
            ws.merge_cells(f'E{start_row+1}:K{start_row+1}')
            
            ws.merge_cells(f'E{start_row+3}:N{start_row+3}')
            ws.merge_cells(f'E{start_row+4}:N{start_row+4}')
            ws.merge_cells(f'E{start_row+5}:N{start_row+5}')
            

            
            ws.merge_cells(f'J{start_row+7+1}:K{start_row+7+1}')
            ws.merge_cells(f'J{start_row+8+1}:K{start_row+8+1}')
            ws.merge_cells(f'L{start_row+7+1}:M{start_row+7+1}')
            ws.merge_cells(f'L{start_row+8+1}:M{start_row+8+1}')

            ws.merge_cells(f'H{start_row+11+1}:J{start_row+11+1}')
            ws.merge_cells(f'L{start_row+11+1}:N{start_row+11+1}')
            ws.merge_cells(f'H{start_row+12+1}:J{start_row+12+1}')
            ws.merge_cells(f'L{start_row+12+1}:N{start_row+12+1}')


            ws.merge_cells(f'E{start_row+14+1}:L{start_row+14+1}')
            ws.merge_cells(f'E{start_row+15+1}:L{start_row+15+1}')
            ws.merge_cells(f'E{start_row+16+1}:L{start_row+16+1}')

            ws.merge_cells(f'E{start_row+18+1}:H{start_row+18+1}')
            ws.merge_cells(f'E{start_row+19+1}:H{start_row+19+1}')
            ws.merge_cells(f'E{start_row+20+1}:H{start_row+20+1}')
            ws.merge_cells(f'E{start_row+21+1}:H{start_row+21+1}')
            ws.merge_cells(f'E{start_row+22+1}:H{start_row+22+1}')
            ws.merge_cells(f'E{start_row+23+1}:H{start_row+23+1}')





            # # Применяем стили к новым строкам
            # for coord, style in styles_to_apply:
            #     ws[coord].style = style

            # for row_num in del_rows:
            #     ws.row_dimensions[row_num].height = 0


            print("render_data")
            print(render_data)
            print("render_data")

            f_name = "АКТ "

            try:f_name += str(render_data["data"]['BS_NUMBER'])
            except: pass

            # try:f_name += " " + render_data["data"]["TYPE_OF_WORK"]
            # except: pass

            try:f_name += " " + render_data["data"]["BS_COMPANY"]
            except: pass

            try:f_name += " " + render_data["data"]["ORDER_REGION"]
            except: pass

            try:f_name += " " + render_data["data"]["BS_ADDRESS"]
            except: pass

            try:f_name = f_name.strip().replace("\\", "-").replace("/", "-")
            except: pass

            try:f_name = f_name.split("\n")[0]
            except: pass


            # Сохраняем созданную копию с измененными значениями
            wb.save(os.path.join(output_folder_path, f_name + ".xlsx"))

            os.remove(os.path.join(output_folder_path, "output.xlsx"))






            try:
                import requests
                import datetime
                def send_report(text=None, process=None, responsible=None):
                    requests.post(f"https://script.google.com/macros/s/AKfycbzDwjE6Pu1a7otho2EHwbI-4yNoEmLijTfwWfI3toWpDpJ6rc-O1pKljV6XMLJmQIyJ/exec?time={datetime.datetime.now().strftime('%d.%m.%Y %H:%M:%S')}&process={process}&responsible={responsible}&text={text}")
                send_report(text="B2B АТП Генератор", process="B2B АТП Генератор", responsible=os.getlogin())
            except: pass

            return {"message": "Файл успешно создан и изменен: " + os.path.join(output_folder_path, os.path.join(output_folder_path, f_name + ".xlsx"))}

        except Exception as e:
            return {"message": f"Произошла ошибка: {str(traceback.format_exc())}"}
    return {"message": f"Нет таблицы"}



# , {{clent}}, {{city}}, {{address}}	
# {{number}}
# {{date}}
# {{budget}} 
# {{budget_with_nds}} 
# {{nds_budget}}


# {{table_start}}
# {{table_end}}



# {{table_number}}
# {{table_work_name}}
# {{table_measure}} 
# {{table_count}}
# {{table_price}}
# {{table_price_with_nds}}


