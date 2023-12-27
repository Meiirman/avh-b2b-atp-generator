import re
import os
import traceback
from openpyxl import load_workbook
from bs4 import BeautifulSoup


def get_TABLE(table):
    TABLE = [   {
        "index":0 ,
        "number":0 ,
        "work_name":0 ,
        "measure":0 ,
        "count":0 ,
        "price":0 ,
        "price_with_nds":0 ,
    }]
    if table:
        rows = table.find_all("tr")
        for index, row in enumerate(rows):
            cells = row.find_all("td")
            row_list = []
            for cell in cells:
                row_list.append(cell.text.strip())
            try:
                i = row_list

                try:
                    TABLE.append(
                        {
                            "index": i[0], # N
                            "number": i[1], # P
                            "work_name": i[2], # D
                            "measure": i[3], # M
                            "count": i[4], # C
                            "price": round(float(str(i[5]).replace(",", ".").replace(" ", ""))/float(str(i[4]).replace(",", ".").replace(" ", "")), 2 ), # T
                            "price_with_nds": round(float(str(i[5]).replace(",", ".").replace(" ", ""))/float(i[4].replace(",", ".").replace(" ", ""))*1.12, 2 ), # S
                        }
                    )
                except: pass
            except:
                print("\n_" * 5)
                print(traceback.format_exc())
                print("\n_" * 5)
                return {"message": "ошибка при получения данных из таблицы html"}
    return TABLE


def replace_p_tags_with_br(html_content):
    html_content = html_content.replace("style=\"font-size:0.12in;\"", "")
    html_content = html_content.replace("align=\"left\"", "")
    html_content = html_content.replace("align=\"center\"", "")
    html_content = html_content.replace("<br >", "")
    html_content = html_content.replace("<br>", "")
    html_content = html_content.replace("<p", "<br")
    html_content = html_content.replace("<b", "<br")
    html_content = html_content.replace("</p>", "")
    html_content = html_content.replace("brr", "br")
    html_content = html_content.replace("</br>", "")
    html_content = html_content.replace("</b>", "")
    html_content = html_content.replace("</b>", "")
    html_content = html_content.replace("<center>", "")
    html_content = html_content.replace("</center>", "")
    html_content = html_content.replace("<br >", "")
    html_content = html_content.replace("<br>", "")
    html_content = html_content.replace("<br/>", "")
    html_content = html_content.replace("brody", "body")
    html_content = html_content.replace("\n", "")
    html_content = html_content.replace("Итого стоимость работ", "\nИтого стоимость работ")
    html_content = html_content.replace("Всего общая стоимость работ", "\nВсего общая стоимость работ")
    html_content = html_content.replace("НДС 12%: ", "\nНДС 12%: ")
    html_content = html_content.replace("Номер заказа:", "\nНомер заказа:")
    html_content = html_content.replace("<", "\n<")
    html_content = html_content.replace("Регион: [", "\nРегион: [<region_code>")
    html_content = html_content.replace("] Номер Заявки", "</region_code>]\n Номер Заявки")
    
    html_content_x = html_content.split("\n")
    for i, e in enumerate(html_content_x):
        if "Итого стоимость работ" in e:
            html_content_x[i] = "<itogo_word>" + e + "</itogo_word>" 
        if "Всего общая стоимость работ" in e:
            html_content_x[i] = "<itogo_total_word>" + e + "</itogo_total_word>" 
        if "НДС 12%: " in e:
            html_content_x[i] = "<NDC_word>" + e + "</NDC_word>" 
        if "к рамочному договору" in e:
            html_content_x[i] = "<get_dogovor_data>" + e + "</get_dogovor_data>" 

    html_content = "\n".join(html_content_x)
    return html_content



def get_data(html_file_path, work_folder, proposal_path):

    html_file = open(html_file_path, "r", encoding="utf-8")  # type: ignore
    html_content = replace_p_tags_with_br(html_file.read())

    # преобразовать html
    soup = BeautifulSoup(html_content, "html.parser")
    body = soup.find("body")

    multi_TABLE = []

    main_tables = body.findChildren(recursive=False)[0].find_all("table")[1:]  # type: ignore
    for i in range(0, len(main_tables), 2):
        multi_TABLE.append(get_TABLE(main_tables[i]))  # DONE
    
    TABLE = multi_TABLE[-1]
    data = {
        "BS_NUMBER" : "", 
        "BS_NAME" : "",
        "BS_COMPANY" : "BS_COMPANYBS_COMPANYBS_COMPANYBS_COMPANY",
        "BS_ADDRESS" : "",
        "ORDER_REGION" : "",
        "ORDER_MANAGER" : "",
        "ORDER_NUMBER" : "",
        "ORDER_DATE" : "",
        "TOTAL_SUMM" : "",
        "TOTAL_NDS" : "",
        "TOTAL_SUMM_NDS" : "",
        "TOTAL_SUMM_NDS_WORD" : "",
        "ORDER_DOGOVOR_NUMBER" : "",
        "ORDER_DOGOVOR_DATE" : "",
        "TABLE" : TABLE,
        "ORDER_MANAGER_POSITION" : "",
        "TYPE_OF_WORK" : "",
    }

    try:      
        proposal_full_path = os.path.join(os.getcwd(), proposal_path)
        wb = load_workbook(proposal_full_path)
        ws = wb.active

        arr = proposal_full_path.split("\\")[-1].split(" ")
        split_word = "qf"
        for word in arr:
            if "crq" in word.lower():
                split_word = word
        try:
            x = proposal_full_path.split("\\")[-1].split(split_word)[1]
            x = x.split(",")[0]
            data["BS_COMPANY"] = x

        except:
            pass

        try: data["ORDER_REGION"] = ws["C11"].value
        except: pass

        try: data["BS_ADDRESS"] = ws["C14"].value
        except: pass

        try: data["BS_NUMBER"] = ws["C3"].value
        except: pass

        try: data["ORDER_DOGOVOR_DATE"] = ws["C4"].value
        except: pass

        try: data["TYPE_OF_WORK"] = ws["C6"].value
        except: pass

        return {"data": data, "message": "Данные успешно прочитаны и записаны в память"}
    except PermissionError:
        return {"message": f"Закройте файл и повторите попытку."}
