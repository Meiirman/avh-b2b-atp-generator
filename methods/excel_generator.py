import json
import os
import traceback
from openpyxl import load_workbook

def get_data(source_path, work_folder, proposal_path) -> dict:
    try:
        full_path = os.path.join(os.getcwd(), source_path)
        workbook = load_workbook(full_path)
        sheet = workbook.active

        # Прочитать данные из столбца D
        column_d_data = [cell.value for cell in sheet['D'][11:] if cell.value is not None]
        # Фильтр: показать только те строки, у которых столбец D не пустой
        filtered_data = [row for row in sheet.iter_rows(min_col=1, max_col=20, min_row=1, values_only=True) if row[3] is not None]
        TABLE = []
        for ix, i in enumerate(filtered_data):
            TABLE.append({
                "index" : str(ix), # №
                "number" : i[0], # №
                "work_name" : i[1], # Наименование работ и оборудования Материалов
                "measure" : i[2], # Ед. измерения
                "count" : i[3], # Кол-во
                "price" : i[4], # Цена без НДС
                "price_with_nds" : i[5], # Цена с НДС              
           })
            
        data = {
            "BS_NUMBER" : "", 
            "BS_NAME" : "",
            "BS_COMPANY" : "BS_COMPANYBS_COMPANYBS_COMPANYBS_COMPANY",
            "BS_ADDRESS" : "",
            "ORDER_REGION" : "",
            "ORDER_MANAGER" : "",
            "ORDER_NUMBER" : "",
            "ORDER_DATE" : "",
            "TOTAL_SUMM" : "",
            "TOTAL_NDS" : "",
            "TOTAL_SUMM_NDS" : "",
            "TOTAL_SUMM_NDS_WORD" : "",
            "ORDER_DOGOVOR_NUMBER" : "",
            "ORDER_DOGOVOR_DATE" : "",
            "TABLE" : TABLE,
            "ORDER_MANAGER_POSITION" : "",
            "TYPE_OF_WORK" : "",
        }

        
        proposal_full_path = os.path.join(os.getcwd(), proposal_path)
        wb = load_workbook(proposal_full_path)
        ws = wb.active

        arr = proposal_full_path.split("\\")[-1].split(" ")
        split_word = "qf"
        for word in arr:
            if "crq" in word.lower():
                split_word = word
        try:
            x = proposal_full_path.split("\\")[-1].split(split_word)[1]
            x = x.split(",")[0]
            data["BS_COMPANY"] = x

        except:
            pass

        try: data["ORDER_REGION"] = ws["C11"].value
        except: pass

        try: data["BS_ADDRESS"] = ws["C14"].value
        except: pass

        try: data["BS_NUMBER"] = ws["C3"].value
        except: pass

        try: data["ORDER_DOGOVOR_DATE"] = ws["C4"].value
        except: pass

        try: data["TYPE_OF_WORK"] = ws["C6"].value
        except: pass

        return {"data": data, "message": "Данные успешно прочитаны и отфильтрованы"}
    except PermissionError:
        return {"message": f"Закройте файл и повторите попытку.\n\nФайл: {full_path}"}


# BS_COMPANY
# ORDER_REGION
# BS_ADDRESS
# BS_NUMBER
# ORDER_DOGOVOR_DATE
# TOTAL_NDS
# TOTAL_SUMM_NDS
    

    